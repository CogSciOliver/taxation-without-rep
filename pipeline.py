# v2.2.2 WORK IN PROGRESS 04.01.2026 13:00, author Danii Oliver 

from __future__ import annotations
import json
from random import sample
import re
from pathlib import Path
from typing import Tuple, List, Dict
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = Path("data")
RULES_PATH = DATA_DIR / "category_rules.json"
VENDOR_MAP_PATH = DATA_DIR / "vendor_map.json"

DEFAULT_CATEGORY = "Uncategorized"

TRANSFER_HINTS = [
    "TRANSFER", "ACH TRANSFER", "ZELLE", "VENMO", "CASH APP", "PAYPAL TRANSFER",
    "INTERNAL TRANSFER", "TO SAVINGS", "FROM SAVINGS",
]
CC_PAYMENT_HINTS = ["CREDIT CARD PAYMENT", "CARD PAYMENT", "AUTOPAY", "ONLINE PAYMENT"]
OWNER_HINTS = ["OWNER DRAW", "OWNER'S DRAW", "OWNER CONTRIBUTION", "CAPITAL CONTRIBUTION"]
LOAN_HINTS = ["LOAN", "LENDING", "PAYMENT TO", "PRINCIPAL", "LINE OF CREDIT"]
PAYROLL_HINTS = ["PAYROLL", "GUSTO", "ADP", "PAYCHEX", "WITHHOLDING", "941", "940"]


def _load_json(path: Path, default):
    try:
        if path.exists():
            return json.loads(path.read_text())
    except Exception:
        pass
    return default


def normalize_bank_csv(df: pd.DataFrame) -> pd.DataFrame:
    def canon(name: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower())

    def find_by_partial(columns, terms):
        ranked = []
        for col in columns:
            c = canon(col)
            score = 0
            for term in terms:
                if term == c:
                    score += 100
                elif term in c:
                    score += 10
            if score > 0:
                ranked.append((score, col))
        ranked.sort(reverse=True, key=lambda x: x[0])
        return ranked[0][1] if ranked else None

    def find_date_by_values(frame: pd.DataFrame, threshold: float = 0.60):
        best_col = None
        best_score = 0.0

        for col in frame.columns:
            sample = frame[col].dropna().astype(str).str.strip()
            sample = sample[sample != ""].head(100)

            if sample.empty:
                continue

            parsed = pd.to_datetime(sample, errors="coerce")
            score = float(parsed.notna().mean())

            if score > best_score:
                best_score = score
                best_col = col

        return best_col if best_score >= threshold else None

    def find_text_fallback(frame: pd.DataFrame, exclude=None):
        exclude = set(exclude or [])
        best_col = None
        best_score = -1

        for col in frame.columns:
            if col in exclude:
                continue

            sample = frame[col].dropna().astype(str).str.strip()
            sample = sample[sample != ""].head(100)

            if sample.empty:
                continue

            c = canon(col)

            if any(term in c for term in ["amount", "debit", "credit", "balance", "account", "number"]):
                continue

            avg_len = sample.str.len().mean()
            uniqueness = sample.nunique(dropna=True)

            score = 0
            if avg_len >= 4:
                score += 1
            if avg_len >= 8:
                score += 1
            score += min(int(uniqueness), 5)

            if score > best_score:
                best_score = score
                best_col = col

        return best_col

    date_col = find_by_partial(df.columns, ["date", "post", "posted", "transaction"])
    if not date_col:
        date_col = find_date_by_values(df)

    primary_desc_col = find_by_partial(
        df.columns,
        ["description", "merchant", "memo", "payee", "name", "detail"]
    )

    fallback_desc_col = find_by_partial(
        df.columns,
        ["classification", "status"]
    )

    def is_usable_text(col, frame):
        sample = frame[col].dropna().astype(str).str.strip()
        sample = sample[sample != ""].head(50)
        return sample.str.len().mean() > 3 if not sample.empty else False

    desc_col = None

    if primary_desc_col and is_usable_text(primary_desc_col, df):
        desc_col = primary_desc_col
    elif fallback_desc_col and is_usable_text(fallback_desc_col, df):
        desc_col = fallback_desc_col
    else:
        desc_col = find_text_fallback(df, exclude=[date_col])

    if not desc_col:
        desc_col = find_text_fallback(df, exclude=[date_col])

    amt_col = find_by_partial(df.columns, ["amount", "amt", "value"])
    debit_col = find_by_partial(df.columns, ["debit", "withdraw", "outflow", "payment"])
    credit_col = find_by_partial(df.columns, ["credit", "deposit", "inflow"])

    if date_col is None:
        raise ValueError("CSV must include a recognizable date column or a column with parseable dates.")

    if desc_col is None:
        raise ValueError("CSV must include a recognizable description column or text column to use as transaction detail.")

    working = df.copy()

    if amt_col is None:
        if debit_col and credit_col:
            working["Amount"] = (
                pd.to_numeric(working[credit_col], errors="coerce").fillna(0)
                - pd.to_numeric(working[debit_col], errors="coerce").fillna(0)
            )
        elif debit_col:
            working["Amount"] = -pd.to_numeric(working[debit_col], errors="coerce")
        elif credit_col:
            working["Amount"] = pd.to_numeric(working[credit_col], errors="coerce")
        else:
            raise ValueError("CSV must include Amount or Debit/Credit columns.")
        
        amt_col = "Amount"

    type_col = find_by_partial(df.columns, ["type"])
      
    amount_series = pd.to_numeric(working[amt_col], errors="coerce").copy()

    if type_col:
        type_text = working[type_col].fillna("").astype(str).str.strip().str.lower()

        expense_mask = type_text.str.contains("expense|debit|withdraw|outflow", regex=True)
        income_mask = type_text.str.contains("income|credit|deposit|inflow", regex=True)

        amount_series = amount_series.where(~expense_mask, -amount_series.abs())
        amount_series = amount_series.where(~income_mask, amount_series.abs())

    out = pd.DataFrame(
        {
            "date": pd.to_datetime(working[date_col], errors="coerce"),
            "description": working[desc_col].astype(str),
            "amount": amount_series,
        }
    )

    out = out.dropna(subset=["date", "amount"]).copy()
    out["description_norm"] = out["description"].str.upper().str.replace(r"\s+", " ", regex=True)
    out["month"] = out["date"].dt.to_period("M").astype(str)
    out["direction"] = out["amount"].apply(lambda x: "in" if x > 0 else "out")
    out["abs_amount"] = out["amount"].abs()

    out["category"] = DEFAULT_CATEGORY
    out["cat_confidence"] = 0.0
    out["cat_source"] = "default"

    out["is_pl_item"] = True
    out["non_pl_reason"] = ""

    return out


def initialize_workspace_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "pl_section" not in df.columns:
        df["pl_section"] = df["amount"].apply(lambda x: "income" if x > 0 else "expense")

    if "cash_flow_bucket" not in df.columns:
        df["cash_flow_bucket"] = ""

    if "source_kind" not in df.columns:
        df["source_kind"] = ""

    return df


def apply_categorization(df: pd.DataFrame) -> pd.DataFrame:
    vendor_map: dict = _load_json(VENDOR_MAP_PATH, {})
    rules: list = _load_json(RULES_PATH, [])

    def categorize(desc_norm):
        desc_text = "" if pd.isna(desc_norm) else str(desc_norm)

        for k, v in vendor_map.items():
            key_text = "" if pd.isna(k) else str(k)
            if key_text and key_text.upper() in desc_text:
                return v, 0.95, "vendor_map"

        for r in rules:
            pat = r.get("pattern")
            cat = r.get("category", DEFAULT_CATEGORY)

            pat_text = "" if pd.isna(pat) else str(pat)
            if pat_text:
                try:
                    if re.search(pat_text, desc_text):
                        return cat, float(r.get("confidence", 0.75)), "rule"
                except re.error:
                    continue

        return DEFAULT_CATEGORY, 0.10, "default"

    df = df.copy()
    df["description_norm"] = df["description_norm"].fillna("").astype(str)

    cats = df["description_norm"].apply(categorize)
    df["category"] = cats.apply(lambda x: x[0])
    df["cat_confidence"] = cats.apply(lambda x: x[1])
    df["cat_source"] = cats.apply(lambda x: x[2])
    return df


def detect_non_pl_items(df: pd.DataFrame) -> pd.DataFrame:
    def mark(mask, reason):
        df.loc[mask, "is_pl_item"] = False
        df.loc[mask, "non_pl_reason"] = reason

    dn = df["description_norm"]

    mark(dn.str.contains("|".join(map(re.escape, TRANSFER_HINTS))), "Likely transfer")
    mark(dn.str.contains("|".join(map(re.escape, CC_PAYMENT_HINTS))), "Likely credit card payment/transfer")
    mark(dn.str.contains("|".join(map(re.escape, OWNER_HINTS))), "Owner draw/contribution (equity)")
    mark(dn.str.contains("|".join(map(re.escape, PAYROLL_HINTS))), "Payroll/withholding (may need payroll filings)")
    mark(dn.str.contains("|".join(map(re.escape, LOAN_HINTS))), "Loan/financing (split principal vs interest)")

    return df


def build_pl_tables(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    pl = df[df["is_pl_item"] == True].copy()
    pl["type"] = pl["amount"].apply(lambda x: "income" if x > 0 else "expense")
    pl["abs_amount"] = pl["amount"].abs()

    by_cat = (
        pl.groupby(["type", "category"], as_index=False)["abs_amount"]
        .sum()
        .rename(columns={"abs_amount": "total"})
        .sort_values(["type", "total"], ascending=[True, False])
    )

    by_month = (
        pl.groupby(["month", "type"], as_index=False)["abs_amount"]
        .sum()
        .rename(columns={"abs_amount": "total"})
        .sort_values(["month", "type"])
    )

    return by_cat, by_month


def build_flags(df: pd.DataFrame) -> pd.DataFrame:
    flags = []
    likely_deductible = {
        "Advertising", "Office", "Supplies", "Utilities", "Legal & Professional",
        "Contract Labor", "Repairs & Maintenance", "Rent/Lease", "Insurance", "Travel", "Meals",
        "Taxes & Licenses", "Interest", "Bank Fees"
    }
    needs_review = {"Meals", "Travel", "Car & Truck", "Other", "Uncategorized"}

    for _, r in df.iterrows():
        if r["amount"] >= 0:
            continue

        if r["is_pl_item"] is False:
            flags.append({
                "date": str(r["date"].date()),
                "description": r["description"],
                "amount": float(r["abs_amount"]),
                "category": r["category"],
                "flag": "Non-P&L candidate",
                "why": r["non_pl_reason"],
            })
            continue

        cat = r["category"]
        if cat in likely_deductible and r["cat_confidence"] >= 0.6:
            flags.append({
                "date": str(r["date"].date()),
                "description": r["description"],
                "amount": float(r["abs_amount"]),
                "category": cat,
                "flag": "Likely deductible",
                "why": f"Category={cat}, confidence={r['cat_confidence']:.2f}",
            })
        elif cat in needs_review or r["cat_confidence"] < 0.6:
            flags.append({
                "date": str(r["date"].date()),
                "description": r["description"],
                "amount": float(r["abs_amount"]),
                "category": cat,
                "flag": "Needs review",
                "why": "Mixed-use/substantiation or low confidence",
            })

    return pd.DataFrame(flags).head(500)


def build_form_checklist(entity_mode: str) -> List[Dict[str, str]]:
    common = [
        {"form": "Bring-to-preparer package", "why": "Attach categorized export + substantiation notes for flagged items."},
        {"form": "1099-K / processor reconciliation", "why": "If you accept card/app payments; reconcile gross receipts vs deposits."},
        {"form": "Mileage / vehicle log", "why": "If you have vehicle costs; keep mileage and business-purpose records."},
        {"form": "Receipts & business purpose notes", "why": "Especially for meals/travel and any mixed-use items."},
    ]

    if entity_mode == "schedule_c":
        return [
            {"form": "Schedule C (Form 1040)", "why": "Sole proprietorship / single-member LLC taxed as disregarded entity."},
            {"form": "Form 4562", "why": "Depreciation/Section 179 for assets; also used for listed property."},
            {"form": "Form 8829", "why": "Home office deduction (if eligible; needs separate inputs)."},
            {"form": "Form 1040-ES", "why": "Estimated taxes if withholding doesn’t cover liability."},
            *common
        ]
    if entity_mode == "partnership_llc":
        return [
            {"form": "Form 1065", "why": "Partnership / multi-member LLC taxed as partnership."},
            {"form": "Schedule K-1 (Form 1065)", "why": "Issued to partners for distributive share reporting."},
            {"form": "Form 4562", "why": "Depreciation/Section 179 as applicable."},
            *common
        ]
    if entity_mode == "s_corp":
        return [
            {"form": "Form 1120-S", "why": "S corporation income and deductions."},
            {"form": "Schedule K-1 (Form 1120-S)", "why": "Issued to shareholders for pass-through items."},
            {"form": "Payroll filings (941/940 etc.)", "why": "If owners/employees are paid wages; coordinate payroll compliance."},
            {"form": "Form 4562", "why": "Depreciation/Section 179 as applicable."},
            *common
        ]
    if entity_mode == "exempt_501c3":
        return [
            {"form": "Form 990 / 990-EZ / 990-N", "why": "Annual exempt organization information return/notice (depends on size)."},
            {"form": "Form 990-T", "why": "If unrelated business taxable income (UBTI) exists."},
            *common
        ]
    if entity_mode == "exempt_501d":
        return [
            {"form": "Form 1065", "why": "§501(d) religious/apostolic orgs generally file Form 1065 to report taxable income allocated to members."},
            {"form": "Schedule K-1 (Form 1065)", "why": "Allocate taxable income to members (members bring K-1 to their preparers)."},
            *common
        ]
    return common


def export_workbook(
    df: pd.DataFrame,
    pl_by_cat: pd.DataFrame,
    pl_by_month: pd.DataFrame,
    flags: pd.DataFrame,
    forms: List[Dict[str, str]],
    entity_mode: str,
    filename: str,
    updates: List[Dict[str, str]],
) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.append(["Source file", filename])
    ws0.append(["Entity mode", entity_mode])
    ws0.append([])
    ws0.append(["Sheets:", "PL_by_Category", "PL_by_Month", "Flags", "Transactions", "Non_PL", "Form_Checklist", "IRS_Updates"])

    def add_df(sheet_name: str, data: pd.DataFrame):
        ws = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(list(r))

    add_df("PL_by_Category", pl_by_cat)
    add_df("PL_by_Month", pl_by_month)
    add_df("Flags", flags)

    tx_cols = ["date", "month", "description", "amount", "category", "cat_confidence", "cat_source", "is_pl_item", "non_pl_reason"]
    add_df("Transactions", df[tx_cols].copy())
    add_df("Non_PL", df[df["is_pl_item"] == False][tx_cols].copy())

    ws_forms = wb.create_sheet("Form_Checklist")
    ws_forms.append(["form", "why"])
    for f in forms:
        ws_forms.append([f.get("form", ""), f.get("why", "")])

    ws_updates = wb.create_sheet("IRS_Updates")
    ws_updates.append(["title", "date", "url", "snippet"])
    for u in updates:
        ws_updates.append([u.get("title", ""), u.get("date", ""), u.get("url", ""), u.get("snippet", "")])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()